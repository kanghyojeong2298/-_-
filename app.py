"""
소포수령증 자동화 웹앱 (Streamlit)
실행: streamlit run app.py
"""
import streamlit as st
import tempfile
import os
import sys
import re
# ── 허용된 직원 이메일 목록 (배포 후 여기에 직원 이메일 추가) ──────────────
ALLOWED_EMAILS = [
    "guwjd2298@gmail.com",
    "help@taxexpert.kr",
    "m0120@taxexpert.kr",
    "m0227@taxexpert.kr",
    "m0429@taxexpert.kr",
    "m0607@taxexpert.kr",
    "m1007@taxexpert.kr",
    "m1211@taxexpert.kr",
    "m1225@taxexpert.kr",
    "m1018@taxexpert.kr",  # 관리자
]
import calendar
from pathlib import Path
from datetime import datetime
sys.path.insert(0, str(Path(__file__).parent))
from modules.pdf_parser   import parse_pdf, detect_pdf_type
from modules.excel_writer import generate_excel, period_labels
CURRENCIES = ["MYR", "PHP", "SGD", "THB", "TWD", "VND", "JPY", "BRL", "MXN"]
CURRENCY_NAMES = {
    "MYR": "말레이시아 링깃 (MYR)", "PHP": "필리핀 페소 (PHP)",
    "SGD": "싱가포르 달러 (SGD)", "THB": "태국 바트 (THB)",
    "TWD": "대만 달러 (TWD)",     "VND": "베트남 동 (VND)",
    "JPY": "일본 엔 (JPY) (100)",
}
# SMBS 통화명 → 코드 매핑
SMBS_NAME_TO_CODE = {
    "MYR": "MYR", "말레이시아": "MYR",
    "PHP": "PHP", "필리핀":    "PHP",
    "SGD": "SGD", "싱가포르":   "SGD",
    "THB": "THB", "태국":      "THB",
    "TWD": "TWD", "대만":      "TWD",
    "VND": "VND", "베트남":     "VND",
    "JPY": "JPY", "일본":      "JPY",
}
DEFAULT_RATES = {
    "MYR": 0.0, "PHP": 0.0, "SGD": 0.0, "THB": 0.0,
    "TWD": 0.0, "VND": 0.0, "JPY": 0.0,
}
# ── 고정환율(내장) 자동 적용 ─────────────────────────────────────────
# data/fixed_rates_2025.json 에 통화별 일별 매매기준율이 들어 있습니다.
# 직원이 환율을 입력할 필요 없이, 소포수령증 발행일에 맞는 환율이 자동 적용됩니다.
# 환율을 갱신하려면 이 JSON 파일만 교체(수정)하면 전체에 반영됩니다.
AUTO_RATE_LABEL = "🔒 자동 (내장 고정환율)"
# 환율 JSON을 여러 위치에서 자동 탐색 (data/ 폴더 안이든, 루트든, 어디 두든 찾음)
_HERE = os.path.dirname(os.path.abspath(__file__))
_FIXED_RATES_CANDIDATES = [
    os.path.join(_HERE, 'data', 'fixed_rates_2025.json'),
    os.path.join(_HERE, 'fixed_rates_2025.json'),
    os.path.join(os.getcwd(), 'data', 'fixed_rates_2025.json'),
    os.path.join(os.getcwd(), 'fixed_rates_2025.json'),
]
def load_fixed_rates() -> dict:
    """내장 JSON에서 통화별 일별 환율을 읽어, 앱 내부 환율 형식(daily 포함)으로 반환."""
    import json
    raw = None
    for _p in _FIXED_RATES_CANDIDATES:
        try:
            with open(_p, encoding='utf-8') as f:
                raw = json.load(f)
            break
        except Exception:
            continue
    if raw is None:
        return {}
    rates_raw = raw.get('rates', {})
    result = {}
    for cur, daymap in rates_raw.items():
        daily = [{'date': d, 'rate': float(r), 'change': 0.0, 'cross': 0.0}
                 for d, r in sorted(daymap.items())]
        if not daily:
            continue
        rs = [d['rate'] for d in daily if d['rate'] > 0]
        result[cur] = {
            'period':        f"{daily[0]['date']} ~ {daily[-1]['date']}",
            'currency':      cur,
            'currency_name': CURRENCY_NAMES.get(cur, cur),
            'average':       round(sum(rs) / len(rs), 2) if rs else 0.0,
            'min':           min(rs) if rs else 0.0,
            'max':           max(rs) if rs else 0.0,
            'min_date': '', 'max_date': '', 'range': 0.0, 'cross_rate': 0.0,
            'daily':         daily,
        }
    return result
# ── 페이지 설정 ─────────────────────────────────────────────────
st.set_page_config(page_title="소포수령증 자동화", page_icon="📦", layout="centered")
# ── 로그인 (Streamlit 네이티브 인증 — Secrets의 [auth] 사용) ──────
_AUTH_ENABLED = False
try:
    _AUTH_ENABLED = "auth" in st.secrets
except Exception:
    _AUTH_ENABLED = False
if _AUTH_ENABLED:
    try:
        _logged_in = st.user.is_logged_in
    except Exception as e:
        st.error(f"로그인 상태 확인 오류: {type(e).__name__}: {e}")
        st.stop()
    if not _logged_in:
        st.markdown(
            """
            <div style="text-align:center; padding:3rem 1rem;">
                <p style="font-size:2rem; font-weight:700; color:#1f4e79;">📦 소포수령증 자동화</p>
                <p style="color:#555; margin-bottom:1.5rem;">사용하려면 Google 계정으로 로그인하세요.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        _lc1, _lc2, _lc3 = st.columns([1, 1.4, 1])
        with _lc2:
            st.button("🔓 Google 계정으로 로그인", type="primary",
                      on_click=st.login, use_container_width=True)
        st.stop()
    _user_email = st.user.get("email", "")
    if ALLOWED_EMAILS and _user_email not in ALLOWED_EMAILS:
        st.error(f"❌ 접근 권한이 없습니다. ({_user_email})\n\n관리자에게 문의하세요.")
        if st.button("로그아웃"):
            st.logout()
        st.stop()
    with st.sidebar:
        _user_name = st.user.get("name", "") or _user_email
        st.markdown(f"**👤 {_user_name}**")
        st.markdown(f"<small>{_user_email}</small>", unsafe_allow_html=True)
        if st.button("로그아웃"):
            st.logout()

st.markdown("""
<style>
    .main-title { font-size:2rem; font-weight:700; color:#1f4e79; margin-bottom:0.2rem; }
    .sub-title  { font-size:1rem; color:#555; margin-bottom:2rem; }
    .warn-box   { background:#fff8e1; border-radius:8px; padding:0.8rem 1.2rem;
                  border-left:4px solid #f9a825; margin-bottom:0.5rem; }
    .info-box   { background:#e3f2fd; border-radius:8px; padding:0.8rem 1.2rem;
                  border-left:4px solid #1565c0; margin-bottom:0.5rem; }
    /* 컬럼 안 작은 버튼(초기화 등): 글자 작게 + 줄바꿈 방지 */
    div[data-testid="stColumn"] .stButton > button,
    div[data-testid="column"]   .stButton > button {
        font-size: 13px !important;
        white-space: nowrap !important;
        padding: 4px 12px !important;
        min-height: 0 !important;
        line-height: 1.3 !important;
    }
    div[data-testid="stColumn"] .stButton > button p,
    div[data-testid="column"]   .stButton > button p,
    div[data-testid="stColumn"] .stButton > button span,
    div[data-testid="column"]   .stButton > button span {
        font-size: 13px !important;
        white-space: nowrap !important;
    }
    /* 사이드바(회색 부분) 너비 축소 — 펼쳤을 때만 적용(접으면 본문 가운데 정렬) */
    section[data-testid="stSidebar"][aria-expanded="true"] {
        width: 200px !important;
        min-width: 200px !important;
    }
    /* 사이드바 로그아웃 버튼: 오른쪽 정렬 + 작게 */
    section[data-testid="stSidebar"] .stButton {
        display: flex !important;
        justify-content: flex-end !important;
    }
    section[data-testid="stSidebar"] .stButton > button {
        font-size: 10px !important;
        padding: 2px 8px !important;
        min-height: 0 !important;
        width: auto !important;
        white-space: nowrap !important;
    }
</style>
""", unsafe_allow_html=True)
st.markdown('<p class="main-title">📦 소포수령증 자동화</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-title">소포수령증 PDF를 업로드하면 매출집계 엑셀을 자동으로 만들어 드립니다.</p>', unsafe_allow_html=True)
st.divider()
# ══════════════════════════════════════════════════════════════════
# STEP 1 — PDF 업로드
# ══════════════════════════════════════════════════════════════════
if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0
st.markdown("### 📄 STEP 1 — 소포수령증 PDF 업로드")
_c_desc, _c_reset = st.columns([6, 1])
_c_desc.caption("쇼피(MY/PH/SG/TH/TW/VN/BR/MX), 라자다 파일을 한꺼번에 올려주세요  \n*큐텐재팬은 STEP2에서 진행해주세요")
if _c_reset.button("🔄 초기화"):
    st.session_state.uploader_key += 1
    st.session_state.qoo10_entries = []
    st.rerun()
uploaded_files = st.file_uploader(
    "PDF 파일 선택 (여러 개 동시 선택 가능)",
    type=["pdf"],
    accept_multiple_files=True,
    label_visibility="collapsed",
    key=f"pdf_uploader_{st.session_state.uploader_key}",
)
if uploaded_files:
    st.markdown("**업로드된 파일:**")
    cols = st.columns(2)
    for i, f in enumerate(uploaded_files):
        pdf_type = detect_pdf_type(f.name)
        icon  = {"shopee":"🛍️","lazada":"🟠","qoo10":"🇯🇵","unknown":"❓"}.get(pdf_type,"📄")
        label = {"shopee":"쇼피","lazada":"라자다","qoo10":"큐텐재팬","unknown":"미확인"}.get(pdf_type,"")
        cols[i%2].markdown(f"{icon} `{f.name}` — {label}")
st.divider()
# ══════════════════════════════════════════════════════════════════
# STEP 2 — 큐텐재팬 정보 입력
# ══════════════════════════════════════════════════════════════════
st.markdown("### 🇯🇵 STEP 2 — 큐텐재팬 정보 입력")
st.markdown(
    '<div class="warn-box">'
    '큐텐재팬 PDF는 이미지 형식이라 <b>자동 추출이 되지 않습니다.</b><br>'
    'PDF를 열어서 아래에 직접 입력해 주세요.'
    '</div>',
    unsafe_allow_html=True,
)
st.write("")
import pandas as pd
def _fmt_date(v: str) -> str:
    """8자리 숫자(20251201)면 2025-12-01로 자동 변환. 그 외는 그대로."""
    d = re.sub(r'\D', '', str(v or ''))
    if len(d) == 8:
        return f"{d[:4]}-{d[4:6]}-{d[6:8]}"
    return str(v or '').strip()
if "qoo10_entries" not in st.session_state:
    st.session_state.qoo10_entries = []
st.caption("한 건씩 입력하고 **➕ 추가**를 누르면 아래 표에 정리됩니다. 추가하면 입력칸(거래기간 포함)이 비워집니다. (발행일이 환율 기준일)")
with st.form("qoo10_add_form", clear_on_submit=True):
    _fp1, _fp2 = st.columns(2)
    _in_ps = _fp1.text_input("거래기간 시작일", placeholder="예: 20251201 (하이픈 자동)")
    _in_pe = _fp2.text_input("거래기간 종료일", placeholder="예: 20251231 (하이픈 자동)")
    _fc1, _fc2, _fc3, _fc4 = st.columns(4)
    _in_amount = _fc1.number_input("금액(JPY)", min_value=0, value=0, format="%d")
    _in_qty    = _fc2.number_input("건수", min_value=0, value=0, format="%d")
    _in_track  = _fc3.text_input("발송번호", placeholder="예: K2512244647017")
    _in_wdate  = _fc4.text_input("발행일", placeholder="예: 20260105 (하이픈 자동)")
    _added = st.form_submit_button("➕ 추가", use_container_width=True)
if _added:
    if _in_amount > 0 or _in_qty > 0 or _in_track.strip():
        st.session_state.qoo10_entries.append({
            "거래기간 시작": _fmt_date(_in_ps),
            "거래기간 종료": _fmt_date(_in_pe),
            "발송번호":  _in_track.strip(),
            "건수":     int(_in_qty),
            "금액(JPY)": float(_in_amount),
            "발행일":   _fmt_date(_in_wdate),
        })
    else:
        st.warning("금액·건수·발송번호 중 하나는 입력해야 합니다.")
if st.session_state.qoo10_entries:
    _df_show = pd.DataFrame(st.session_state.qoo10_entries)
    _df_show.index = range(1, len(_df_show) + 1)
    _df_show["금액(JPY)"] = _df_show["금액(JPY)"].map(lambda x: f"{int(x):,}")
    _df_show["건수"] = _df_show["건수"].map(lambda x: f"{int(x):,}")
    st.table(_df_show)
    _tot_amt = sum(e["금액(JPY)"] for e in st.session_state.qoo10_entries)
    _tot_qty = sum(e["건수"] for e in st.session_state.qoo10_entries)
    st.caption(f"합계: {len(st.session_state.qoo10_entries)}건 / 수량 {int(_tot_qty):,} / 금액 {int(_tot_amt):,} JPY")
    if st.button("🗑️ 전체 삭제"):
        st.session_state.qoo10_entries = []
        st.rerun()
else:
    st.caption("아직 추가된 건이 없습니다.")
st.divider()
# ══════════════════════════════════════════════════════════════════
# STEP 3 — 환율 입력
# ══════════════════════════════════════════════════════════════════
st.markdown("### 💱 STEP 3 — 환율")
st.markdown(
    "📌 **소포수령증 발행일(작성일자)** 기준 환율을 적용합니다 (주말·공휴일이면 직전 영업일)  \n"
    "👉 환율 조회: [서울외국환중개(SMBS)](http://www.smbs.biz/ExRate/StdExRate.jsp)"
)
st.write("")
rate_mode = st.radio(
    "환율 입력 방식",
    [AUTO_RATE_LABEL, "📊 SMBS 엑셀 업로드", "✏️ 직접 입력"],
    horizontal=True,
    label_visibility="collapsed",
)
manual_rates = {cur: 0.0 for cur in CURRENCIES}
smbs_excel_files = []
if rate_mode == AUTO_RATE_LABEL:
    _fr = load_fixed_rates()
    if _fr:
        _curs   = ", ".join(_fr.keys())
        _missing = [c for c in CURRENCIES if c not in _fr]
        _period  = next(iter(_fr.values())).get('period', '')
        st.markdown(
            '<div class="info-box">'
            '✅ <b>환율 자동 적용</b><br>'
            '환율을 직접 입력하지 않아도 됩니다.<br>'
            '소포수령증 <b>발행일에 맞는 날짜의 환율</b>로 자동 적용됩니다.<br>'
            f'<small>내장 환율 기간: {_period}<br>적용 통화: {_curs}</small>'
            '</div>',
            unsafe_allow_html=True,
        )
        if _missing:
            st.caption(f"⚠️ 내장 환율 없는 통화: {', '.join(_missing)} → 0 처리 (필요 시 다른 모드 사용)")
    else:
        st.warning(
            "⚠️ 내장 환율 파일(`data/fixed_rates_2025.json`)을 찾을 수 없습니다.  \n"
            "다른 입력 방식을 선택하거나, 데이터 파일을 추가해 주세요."
        )
elif rate_mode == "📊 SMBS 엑셀 업로드":
    st.markdown(
        '<div class="info-box">'
        '<b>SMBS 엑셀 다운로드 방법:</b><br>'
        '1. <a href="http://www.smbs.biz/ExRate/StdExRate.jsp" target="_blank">SMBS 사이트</a> 접속<br>'
        '2. 통화 선택 → 기간 설정 → 조회 → <b>엑셀 저장</b> 클릭<br><br>'
        '<b>📁 업로드 방식 (둘 다 지원)</b><br>'
        '• <b>방식 A</b> — 통화별 엑셀 파일을 따로 저장 후 <b>여러 파일 동시 업로드</b><br>'
        '• <b>방식 B</b> — 엑셀 하나에 <b>여러 시트(탭)</b>로 통화별 데이터 정리 후 업로드<br>'
        '&nbsp;&nbsp;&nbsp;(시트 이름에 통화코드 포함 권장: 예 <i>MYR, JPY, SGD</i>…)'
        '</div>',
        unsafe_allow_html=True,
    )
    smbs_excel_files = st.file_uploader(
        "SMBS 엑셀 파일 업로드",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        key="smbs_excel",
        label_visibility="collapsed",
    )
    if smbs_excel_files:
        st.caption(f"✅ {len(smbs_excel_files)}개 파일 업로드됨")
else:
    st.caption("거래기간 마지막날 환율을 직접 입력하세요 (0이면 해당 통화 미사용)")
    c1, c2, c3, c4 = st.columns(4)
    manual_rates['MYR'] = c1.number_input("🇲🇾 MYR",        value=0.0, format="%.2f")
    manual_rates['PHP'] = c2.number_input("🇵🇭 PHP",        value=0.0, format="%.2f")
    manual_rates['SGD'] = c3.number_input("🇸🇬 SGD",        value=0.0, format="%.2f")
    manual_rates['THB'] = c4.number_input("🇹🇭 THB",        value=0.0, format="%.2f")
    c5, c6, c7, _ = st.columns(4)
    manual_rates['TWD'] = c5.number_input("🇹🇼 TWD",        value=0.0, format="%.2f")
    manual_rates['VND'] = c6.number_input("🇻🇳 VND",        value=0.0, format="%.2f")
    manual_rates['JPY'] = c7.number_input("🇯🇵 JPY (100엔)", value=0.0, format="%.2f")
st.divider()
# ══════════════════════════════════════════════════════════════════
# STEP 4 — 처리 시작
# ══════════════════════════════════════════════════════════════════
st.markdown("### ⚡ STEP 4 — 처리 시작")
process_btn = st.button(
    "🚀 엑셀 파일 생성하기",
    type="primary",
    use_container_width=True,
    disabled=not bool(uploaded_files),
)
if not uploaded_files:
    st.caption("⬆️ PDF 파일을 먼저 업로드해 주세요")
# ── SMBS 엑셀 파싱 함수 ──────────────────────────────────────────
def _parse_smbs_sheet(ws, fallback_currency=None) -> tuple:
    """
    단일 시트에서 날짜/환율 데이터 추출.
    반환: (currency_code, daily_list) 또는 (None, [])
    """
    daily = []
    currency = None
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        cell0 = str(row[0]).strip()
        # 날짜 행 감지: YYYY.MM.DD 또는 YYYY-MM-DD
        date_match = re.match(r'(\d{4})[.\-](\d{2})[.\-](\d{2})', cell0)
        if date_match:
            date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
            try:
                rate   = float(str(row[2]).replace(',', '')) if row[2] is not None else 0.0
                change = float(str(row[3]).replace(',', '')) if row[3] is not None else 0.0
                cross  = float(str(row[4]).replace(',', '')) if row[4] is not None else 0.0
            except (ValueError, TypeError):
                continue
            # 통화 코드 추출 (두 번째 열: 통화명)
            if currency is None and row[1]:
                cur_name = str(row[1])
                for key, code in SMBS_NAME_TO_CODE.items():
                    if key in cur_name:
                        currency = code
                        break
            daily.append({'date': date_str, 'rate': rate,
                          'change': change, 'cross': cross})
        else:
            # 헤더/통화명 행에서 통화 코드 추출 시도
            if currency is None:
                for val in row:
                    if val is None:
                        continue
                    val_str = str(val)
                    for key, code in SMBS_NAME_TO_CODE.items():
                        if key in val_str:
                            currency = code
                            break
                    if currency:
                        break
    # 시트 이름에서 통화 코드 추출 시도
    if currency is None and fallback_currency:
        fname_upper = fallback_currency.upper()
        for code in CURRENCIES:
            if code in fname_upper:
                currency = code
                break
    return currency, daily
def _build_currency_entry(currency, daily):
    """daily 리스트로 통화 환율 딕셔너리 생성"""
    rates_only = [d['rate'] for d in daily if d['rate'] > 0]
    avg = round(sum(rates_only) / len(rates_only), 2) if rates_only else 0.0
    return {
        'period':        f"{daily[0]['date']} ~ {daily[-1]['date']}",
        'currency':      currency,
        'currency_name': CURRENCY_NAMES.get(currency, currency),
        'average':       avg,
        'min':           min(rates_only) if rates_only else 0.0,
        'max':           max(rates_only) if rates_only else 0.0,
        'min_date':      '', 'max_date': '',
        'range':         0.0, 'cross_rate': 0.0,
        'daily':         daily,
    }
def parse_smbs_excel_files(excel_files) -> dict:
    """
    SMBS에서 다운로드한 엑셀 파일들을 파싱하여 통화별 일별 환율 딕셔너리 반환.
    지원 형식:
    1. 통화별 별도 파일 (각 파일에 시트 1개)
    2. 단일 파일에 여러 시트 (시트마다 통화 1개)
    반환: {'MYR': {'period':..., 'currency':'MYR', 'average':..., 'daily':[...]}, ...}
    """
    import openpyxl
    from io import BytesIO
    result = {}
    for ef in excel_files:
        try:
            file_bytes = ef.read()
            wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
            if len(wb.sheetnames) > 1:
                # ── 다중 시트: 시트마다 통화 1개 ──
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    currency, daily = _parse_smbs_sheet(ws, fallback_currency=sheet_name)
                    # 시트명에서 통화 코드 직접 추출 (예: "MYR", "JPY 환율" 등)
                    if currency is None:
                        sn_upper = sheet_name.upper()
                        for code in CURRENCIES:
                            if code in sn_upper:
                                currency = code
                                break
                    if daily and currency:
                        result[currency] = _build_currency_entry(currency, daily)
                    elif daily and not currency:
                        st.caption(f"  ⚠️ `{ef.name}` → 시트 '{sheet_name}' 통화 코드 인식 불가 (건너뜀)")
            else:
                # ── 단일 시트: 파일 1개 = 통화 1개 ──
                ws = wb.active
                currency, daily = _parse_smbs_sheet(ws, fallback_currency=ef.name)
                # 파일명에서 통화 코드 추출 시도
                if currency is None:
                    fname_upper = ef.name.upper()
                    for code in CURRENCIES:
                        if code in fname_upper:
                            currency = code
                            break
                if daily and currency:
                    result[currency] = _build_currency_entry(currency, daily)
                elif daily and not currency:
                    st.caption(f"  ⚠️ `{ef.name}` 통화 코드 인식 불가 (건너뜀)")
        except Exception as e:
            st.warning(f"⚠️ {ef.name} 파싱 오류: {e}")
    return result
def _build_rate_dict(cur, avg, start='', end=''):
    """수동 입력값으로 환율 딕셔너리 생성 (daily 없음)"""
    return {
        'period':        f"{start} ~ {end}",
        'currency':      cur,
        'currency_name': CURRENCY_NAMES.get(cur, cur),
        'average':       avg,
        'min': avg, 'min_date': '', 'max': avg, 'max_date': '',
        'range': 0.0, 'cross_rate': 0.0, 'daily': [],
    }
# ── 처리 실행 ────────────────────────────────────────────────────
if process_btn and uploaded_files:
    progress_bar = st.progress(0, text="처리 시작...")
    status_area  = st.empty()
    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir = Path(tmpdir)
            # ── PDF 임시 저장 ──
            progress_bar.progress(10, text="📥 PDF 저장 중...")
            pdf_paths = []
            for f in uploaded_files:
                dest = tmpdir / f.name
                dest.write_bytes(f.read())
                pdf_paths.append(dest)
            # ── PDF 파싱 ──
            progress_bar.progress(25, text="📄 PDF 파싱 중...")
            shopee_results = []
            lazada_result  = None
            qoo10_result   = None
            parse_log      = []
            for pdf_path in pdf_paths:
                pdf_type_hint = detect_pdf_type(str(pdf_path))
                result = parse_pdf(str(pdf_path))
                if result is None:
                    if pdf_type_hint == 'qoo10':
                        parse_log.append(
                            f"⚠️ `{pdf_path.name}` → 큐텐 자동 추출 불가 "
                            f"(STEP 2에서 직접 입력해주세요)"
                        )
                    else:
                        parse_log.append(f"⚠️ `{pdf_path.name}` — 파싱 실패")
                    continue
                ptype = result.get("type")
                if ptype == "shopee":
                    shopee_results.append(result)
                    parse_log.append(
                        f"✅ `{pdf_path.name}` → 쇼피 {result.get('currency','')} "
                        f"{result.get('total_qty',0):,}건 / {result.get('total_amount',0):,.2f} "
                        f"[{result.get('period_start','')}~{result.get('period_end','')}]"
                    )
                elif ptype == "lazada":
                    lazada_result = result
                    n = len(result.get("items", []))
                    parse_log.append(
                        f"✅ `{pdf_path.name}` → 라자다 {n}개국 "
                        f"[{result.get('period_start','')}~{result.get('period_end','')}]"
                    )
                elif ptype == "qoo10":
                    qoo10_result = result
                    parse_log.append(
                        f"✅ `{pdf_path.name}` → 큐텐 OCR 성공 "
                        f"{result.get('qty',0):,}건 / JPY {result.get('amount',0):,.0f} "
                        f"[{result.get('period_start','')}~{result.get('period_end','')}]"
                    )
            # ── 큐텐 수동 입력 적용 (추가된 표) ──
            qoo10_entries = []
            for _row in st.session_state.get("qoo10_entries", []):
                _amt = float(_row.get("금액(JPY)", 0) or 0)
                _qty = int(_row.get("건수", 0) or 0)
                _trk = str(_row.get("발송번호", "") or "").strip()
                _wdt = str(_row.get("발행일", "") or "").strip()
                if _amt > 0 or _qty > 0 or _trk:
                    qoo10_entries.append({
                        "tracking_no": _trk, "qty": _qty,
                        "amount": _amt, "write_date": _wdt,
                        "period_start": str(_row.get("거래기간 시작", "") or "").strip(),
                        "period_end": str(_row.get("거래기간 종료", "") or "").strip(),
                    })
            if qoo10_entries:
                _total_amt = sum(e["amount"] for e in qoo10_entries)
                _total_qty = sum(e["qty"] for e in qoo10_entries)
                _first_wd  = next((e["write_date"] for e in qoo10_entries if e["write_date"]), "")
                base = qoo10_result or {
                    "type": "qoo10", "carrier": "국제로지스틱",
                    "destination": "JP", "currency": "JPY",
                }
                _ps = [r.get("거래기간 시작", "") for r in st.session_state.get("qoo10_entries", []) if r.get("거래기간 시작")]
                _pe = [r.get("거래기간 종료", "") for r in st.session_state.get("qoo10_entries", []) if r.get("거래기간 종료")]
                base['period_start'] = (min(_ps) if _ps else '') or base.get('period_start', '')
                base['period_end']   = (max(_pe) if _pe else '') or base.get('period_end', '')
                base['write_date']   = _first_wd or base.get('write_date', '')
                base['tracking_no']  = qoo10_entries[0]["tracking_no"]
                base['amount']       = _total_amt
                base['qty']          = _total_qty
                base['entries']      = qoo10_entries
                qoo10_result = base
                parse_log.append(
                    f"📝 큐텐 수동 입력 — {len(qoo10_entries)}건 입력 / 합계 {_total_qty:,}건 / {int(_total_amt):,} JPY"
                )
            # 파싱 결과 표시
            with status_area.container():
                st.markdown("**📋 파싱 결과**")
                for log in parse_log:
                    st.markdown(log)
            # ── 거래기간 파악 ──
            all_starts, all_ends = [], []
            for sd in shopee_results:
                if sd.get('period_start'): all_starts.append(sd['period_start'])
                if sd.get('period_end'):   all_ends.append(sd['period_end'])
            if lazada_result:
                if lazada_result.get('period_start'): all_starts.append(lazada_result['period_start'])
                if lazada_result.get('period_end'):   all_ends.append(lazada_result['period_end'])
            if qoo10_result:
                if qoo10_result.get('period_start'): all_starts.append(qoo10_result['period_start'])
                if qoo10_result.get('period_end'):   all_ends.append(qoo10_result['period_end'])
            # 연월 추정
            year, month = None, None
            for p in pdf_paths:
                m = re.search(r"(\d{4})(\d{2})\d{2}", p.name)
                if m:
                    year, month = int(m.group(1)), int(m.group(2))
                    break
            if not year:
                today = datetime.today()
                month = today.month - 1 or 12
                year  = today.year if today.month > 1 else today.year - 1
            fetch_start = min(all_starts) if all_starts else f'{year}-{month:02d}-01'
            fetch_end   = max(all_ends)   if all_ends   else f'{year}-{month:02d}-{calendar.monthrange(year,month)[1]:02d}'
            # ── 환율 처리 ──
            progress_bar.progress(50, text="💱 환율 처리 중...")
            rates = {}
            # 환율 소스 결정: 자동(내장) → SMBS 업로드 → (둘 다 아니면) 직접 입력
            rate_source = None
            if rate_mode == AUTO_RATE_LABEL:
                rate_source = load_fixed_rates()
            elif rate_mode == "📊 SMBS 엑셀 업로드" and smbs_excel_files:
                rate_source = parse_smbs_excel_files(smbs_excel_files)
            if rate_source is not None:
                from modules.exchange_rate import get_rate_for_date, avg_rate_for_period
                rate_log = []
                missing  = []
                for cur in CURRENCIES:
                    if cur in rate_source and rate_source[cur].get('daily'):
                        rates[cur] = rate_source[cur]
                        _notes = []
                        # 큐텐(JPY): 거래기간 평균환율
                        if cur == 'JPY' and qoo10_result:
                            _qs = qoo10_result.get('period_start', '')
                            _qe = qoo10_result.get('period_end', '')
                            _r = avg_rate_for_period(rate_source[cur], _qs, _qe)
                            _notes.append(f"큐텐 거래기간 평균 **{_r:.2f}**")
                        # 쇼피: 건별 발행일자 환율
                        if any(s.get('currency') == cur for s in shopee_results):
                            _notes.append("쇼피 발행일별 환율")
                        # 라자다: 거래기간 평균환율
                        if lazada_result and any(it.get('currency') == cur for it in lazada_result.get('items', [])):
                            _lps = lazada_result.get('period_start', '')
                            _lpe = lazada_result.get('period_end', '')
                            _lr = avg_rate_for_period(rate_source[cur], _lps, _lpe)
                            _notes.append(f"라자다 거래기간 평균 **{_lr:.2f}**")
                        if _notes:
                            rate_log.append(f"✅ {cur}: " + " / ".join(_notes))
                    else:
                        rates[cur] = _build_rate_dict(cur, 0.0, fetch_start, fetch_end)
                        rate_log.append(f"⚠️ {cur}: 환율 데이터 없음 → 0")
                        missing.append(cur)
                if missing:
                    with status_area.container():
                        st.markdown("**📋 파싱 결과**")
                        for log in parse_log:
                            st.markdown(log)
                        st.warning(f"환율 데이터 없는 통화: **{', '.join(missing)}** → 0으로 처리됩니다.")
            else:
                # 직접 입력
                rates    = {cur: _build_rate_dict(cur, manual_rates.get(cur, 0.0), fetch_start, fetch_end)
                            for cur in CURRENCIES}
                rate_log = [f"✏️ {cur}: **{manual_rates.get(cur,0.0):.2f}**" for cur in CURRENCIES]
            progress_bar.progress(75, text="📊 엑셀 생성 중...")
            # ── 엑셀 생성 ──
            _disp_lbl, _fname_lbl = period_labels(shopee_results, lazada_result, qoo10_result,
                                                  fallback=f'{year}년 {month:02d}월')
            _fsafe = _fname_lbl
            for _ch in '\\/:*?"<>|':
                _fsafe = _fsafe.replace(_ch, ',')
            if not _fsafe:
                _fsafe = f'{year}{month:02d}'
            output_path = tmpdir / f"매출집계_{_fsafe}.xlsx"
            generate_excel(
                shopee_results=shopee_results,
                lazada_result=lazada_result,
                qoo10_result=qoo10_result,
                rates=rates,
                output_path=str(output_path),
                year=year,
                month=month,
            )
            progress_bar.progress(100, text="✅ 완료!")
            excel_bytes = output_path.read_bytes()
        # ── 결과 ──
        st.success(f"✅ 엑셀 생성 완료! — {_disp_lbl}")
        with st.expander("💱 적용된 환율 (쇼피=발행일별 · 라자다/큐텐=거래기간 평균)"):
            for log in rate_log:
                st.markdown(log)
        st.download_button(
            label=f"⬇️  매출집계_{_fsafe}.xlsx  다운로드",
            data=excel_bytes,
            file_name=f"매출집계_{_fsafe}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)
        with st.expander("📋 생성된 시트 목록"):
            cols = st.columns(3)
            for i, name in enumerate(wb.sheetnames):
                cols[i%3].markdown(f"• {name}")
    except Exception as e:
        progress_bar.progress(100, text="오류 발생")
        st.error(f"❌ 오류: {e}")
        st.exception(e)
# ── 하단 안내 ────────────────────────────────────────────────────
st.divider()
with st.expander("📌 파일명 규칙 안내"):
    st.markdown("""
| 파일명 패턴 | 플랫폼 |
|---|---|
| `유엠(UM)_MY_*.pdf` | 쇼피 말레이시아 |
| `유엠(UM)_PH_*.pdf` | 쇼피 필리핀 |
| `유엠(UM)_SG_*.pdf` | 쇼피 싱가폴 |
| `유엠(UM)_TH_*.pdf` | 쇼피 태국 |
| `유엠(UM)_TW_*.pdf` | 쇼피 대만 |
| `유엠(UM)_VN_*.pdf` | 쇼피 베트남 |
| `라자다_*.pdf` | 라자다 |
| `큐텐재팬_*.pdf` | 큐텐재팬 (STEP 2에서 수동 입력) |
""")
