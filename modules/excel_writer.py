"""
엑셀 생성 모듈 — 기존 매출집계 파일과 동일한 형식으로 출력
환율 적용 기준: 소포수령증 발행일(write_date) 환율
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from typing import Optional
import re
from pathlib import Path


# ── 스타일 정의 ────────────────────────────────────────────────
HEADER_FILL   = PatternFill('solid', start_color='D9E1F2', end_color='D9E1F2')
SUBHEAD_FILL  = PatternFill('solid', start_color='E2EFDA', end_color='E2EFDA')
GRAY_FILL     = PatternFill('solid', start_color='F2F2F2', end_color='F2F2F2')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

FONT_DEFAULT = Font(name='맑은 고딕', size=9)
FONT_BOLD    = Font(name='맑은 고딕', size=9, bold=True)
FONT_TITLE   = Font(name='맑은 고딕', size=11, bold=True)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right',  vertical='center')

# SMBS 환율은 일부 통화가 100단위 기준 (원화 = 외화 × 환율 / 제수)
RATE_DIVISOR = {
    'JPY': 100,   # 100엔 기준
    'VND': 100,   # 100동 기준
}


def _style(cell, font=None, fill=None, align=None, border=None, num_format=None):
    if font:      cell.font       = font
    if fill:      cell.fill       = fill
    if align:     cell.alignment  = align
    if border:    cell.border     = border
    if num_format: cell.number_format = num_format


# ── 소포수령증 표 열 그룹 (값 열 + 사이 빈 열을 병합해 깔끔하게 이어줌) ──
_RECEIPT_GROUPS_2 = [(1, 3), (4, 6), (7, 10), (11, 12), (13, 15), (16, 19)]
_RECEIPT_GROUPS_3 = [(1, 3), (4, 6), (7, 10), (11, 12), (13, 15), (16, 17), (18, 18), (19, 19)]


def _merge_row(ws, row, groups, border=None):
    """한 행에서 각 열 그룹을 병합하고(2칸 이상), 그룹 전체에 테두리를 적용."""
    for c1, c2 in groups:
        if c2 > c1:
            ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        if border is not None:
            for col in range(c1, c2 + 1):
                ws.cell(row=row, column=col).border = border


def _get_rate(rates: dict, currency: str, date_str: str) -> float:
    """
    발행일(write_date) 기준 환율 반환.
    daily 데이터 없으면 average(수동입력값) 반환.
    date_str이 비어 있으면 average 반환.
    """
    from .exchange_rate import get_rate_for_date
    rate_data = rates.get(currency)
    if not rate_data:
        return 0.0
    # daily 데이터가 없으면 average 반환 (수동입력 모드)
    if not rate_data.get('daily'):
        return rate_data.get('average', 0.0)
    if not date_str:
        return rate_data.get('average', 0.0)
    rate = get_rate_for_date(rate_data, date_str)
    if rate == 0.0:
        rate = rate_data.get('average', 0.0)
    return rate


# ── 환율 시트 작성 ──────────────────────────────────────────────

def write_exchange_rate_sheet(ws, rate_data: dict):
    """환율(XXX) 시트를 SMBS 데이터로 채움"""
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12

    if rate_data is None:
        ws['A1'] = '환율 데이터 없음 (수동 입력 필요)'
        return

    # 제목
    ws['A1'] = '기간별 매매기준율'
    _style(ws['A1'], font=FONT_BOLD)
    ws['A2'] = f"기간 : {rate_data['period']}"

    # 평균환율 통계
    ws['A4'] = '평균환율'
    _style(ws['A4'], font=FONT_BOLD)

    headers5 = ['평균환율', '최저치', '기록일', '최고치', '기록일', '등락폭', 'Cross Rate']
    for col, h in enumerate(headers5, 1):
        c = ws.cell(row=5, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    vals6 = [
        rate_data.get('average', ''),
        rate_data.get('min', ''),
        rate_data.get('min_date', ''),
        rate_data.get('max', ''),
        rate_data.get('max_date', ''),
        rate_data.get('range', ''),
        rate_data.get('cross_rate', ''),
    ]
    for col, v in enumerate(vals6, 1):
        c = ws.cell(row=6, column=col, value=v)
        _style(c, font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER)

    # 일별
    ws['A7'] = '일별 매매기준율'
    _style(ws['A7'], font=FONT_BOLD)

    headers9 = ['날짜', '통화명', '환율', '전일대비', 'Cross Rate']
    for col, h in enumerate(headers9, 1):
        c = ws.cell(row=9, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    for r, d in enumerate(rate_data.get('daily', []), 10):
        vals = [d['date'], rate_data.get('currency_name', ''), d['rate'], d['change'], d['cross']]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            _style(c, font=FONT_DEFAULT, align=CENTER if col != 2 else LEFT, border=THIN_BORDER)


# ── 쇼피 소포수령증 시트 작성 ───────────────────────────────────

def write_shopee_sheet(ws, shopee_data: dict, rates: dict) -> int:
    """
    쇼피(MYR) 등 국가별 쇼피 시트 작성
    환율: 소포수령증 발행일(write_date) 기준
    """
    col_widths = [16, 3, 3, 12, 3, 3, 20, 3, 3, 3, 5, 3, 5, 3, 3, 12, 3, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    currency   = shopee_data.get('currency', '')
    carrier    = shopee_data.get('carrier', '주)두라로지스틱스')
    country    = shopee_data.get('country', '')
    period_end = shopee_data.get('period_end', '')
    divisor    = RATE_DIVISOR.get(currency, 1)   # VND·JPY → 100, 나머지 → 1

    # ── 행 1: 제목 헤더 ──
    ws.merge_cells('A1:J1')
    ws['A1'] = (
        f"해외배송 소포 수령증\n"
        f"Registration No. 117-81-45551\n"
        f"529-12-02268\n맹진열"
    )
    _style(ws['A1'], font=FONT_BOLD, align=CENTER)
    ws.row_dimensions[1].height = 55

    ws.merge_cells('L1:S1')
    ws['L1'] = (
        f"해외배송기간: {shopee_data.get('period_start','')} ~ {period_end}\n"
        f"유엠(UM)               서울특별시 광진구\n"
        f"광나루로 556, 1동 2층"
    )
    _style(ws['L1'], font=FONT_DEFAULT, align=LEFT)

    # ── 행 2-4: 인적사항 라벨 ──
    ws['A2'] = '사업자등록번호\n대표자 성명 거래기간'
    ws['A3'] = '상호(법인명) 작성일자'
    ws['A4'] = '사업장소재지'
    for row in [2, 3, 4]:
        _style(ws.cell(row=row, column=1), font=FONT_DEFAULT, align=LEFT)

    # ── 행 5: 거래기간, 작성일자 ──
    ws['A5'] = f"{shopee_data.get('period_start','')} ~ {period_end}"
    ws['I5'] = shopee_data.get('write_date', '')
    _style(ws['A5'], font=FONT_DEFAULT)
    _style(ws['I5'], font=FONT_DEFAULT)

    # ── 행 6: 섹션 2 제목 ──
    ws.merge_cells('A6:S6')
    ws['A6'] = '2. 해외배송 소포 수령 수량'
    _style(ws['A6'], font=FONT_BOLD, fill=SUBHEAD_FILL)

    # ── 행 7: 헤더 ──
    headers7 = {
        'A': '해외배송업체', 'D': '배송국가', 'G': '기간', 'K': '통화', 'M': '발송수량', 'P': '발송금액'
    }
    for col_letter, val in headers7.items():
        c = ws[f'{col_letter}7']
        c.value = val
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)
    _merge_row(ws, 7, _RECEIPT_GROUPS_2, border=THIN_BORDER)
    for col_letter in headers7:
        _style(ws[f'{col_letter}7'], font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    # ── 행 8: 데이터 ──
    ws['A8'] = carrier
    ws['D8'] = country
    ws['G8'] = f"{shopee_data.get('period_start','')} ~ {period_end}"
    ws['K8'] = currency
    ws['M8'] = shopee_data.get('total_qty', 0)
    ws['P8'] = shopee_data.get('total_amount', 0.0)
    _merge_row(ws, 8, _RECEIPT_GROUPS_2, border=THIN_BORDER)
    for col in ['A', 'D', 'G', 'K', 'M', 'P']:
        _style(ws[f'{col}8'], font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER)

    # ── 행 9: 합계 ──
    ws['M9'] = shopee_data.get('total_qty', 0)
    ws['P9'] = shopee_data.get('total_amount', 0.0)
    ws['G9'] = '합계'
    _merge_row(ws, 9, _RECEIPT_GROUPS_2, border=THIN_BORDER)
    _style(ws['G9'], font=FONT_BOLD, align=CENTER, border=THIN_BORDER)
    _style(ws['M9'], font=FONT_BOLD, align=CENTER, border=THIN_BORDER)
    _style(ws['P9'], font=FONT_BOLD, align=CENTER, border=THIN_BORDER)

    # ── 행 10: 섹션 3 제목 ──
    ws.merge_cells('A10:O10')
    ws['A10'] = '3. 해외배송 내역'
    _style(ws['A10'], font=FONT_BOLD, fill=SUBHEAD_FILL)

    # ── 행 11: 컬럼 헤더 ──
    col_headers = {
        'A': '해외배송업체', 'D': '발행일', 'G': '운송장번호',
        'K': '도착국가', 'M': '발송수량', 'P': '수출신고금액', 'R': '환율', 'S': '원화'
    }
    for col_letter, val in col_headers.items():
        c = ws[f'{col_letter}11']
        c.value = val
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)
    _merge_row(ws, 11, _RECEIPT_GROUPS_3, border=THIN_BORDER)
    for col_letter in col_headers:
        _style(ws[f'{col_letter}11'], font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    # ── 행 12+: 거래 데이터 (각 행의 발행일 기준 환율 개별 적용) ──
    row = 12
    total_krw = 0
    for tx in shopee_data.get('transactions', []):
        # 해당 거래의 발행일(tx['date']) 기준 환율 개별 조회
        tx_rate = _get_rate(rates, currency, tx['date'])
        krw = round(tx['amount'] * tx_rate / divisor)
        total_krw += krw

        ws.cell(row=row, column=1,  value=tx['carrier'])
        ws.cell(row=row, column=4,  value=tx['date'])
        ws.cell(row=row, column=7,  value=tx['tracking_no'])
        ws.cell(row=row, column=11, value=tx['country'])
        ws.cell(row=row, column=13, value=tx['qty'])
        ws.cell(row=row, column=16, value=tx['amount'])
        ws.cell(row=row, column=18, value=tx_rate)
        ws.cell(row=row, column=19, value=krw)

        _merge_row(ws, row, _RECEIPT_GROUPS_3, border=THIN_BORDER)
        for col in [1, 4, 7, 11, 13, 16, 18, 19]:
            c = ws.cell(row=row, column=col)
            _style(c, font=FONT_DEFAULT, align=CENTER if col != 1 else LEFT, border=THIN_BORDER)

        row += 1

    # 원화 합계
    ws['S10'] = total_krw
    ws['Q10'] = shopee_data.get('total_amount', 0.0)

    # ── 푸터 ──
    footer_row = row + 1
    ws.merge_cells(f'A{footer_row}:S{footer_row}')
    ws[f'A{footer_row}'] = '상기 내역은 판매자가 두라로지스틱스를 통하여 해외 배송한 내역임을 증명합니다'
    _style(ws[f'A{footer_row}'], font=FONT_DEFAULT, align=CENTER)

    footer_row += 1
    ws[f'A{footer_row}'] = '상호 (법인명)'
    ws[f'C{footer_row}'] = '두라로지스틱스'
    ws[f'H{footer_row}'] = '사업자 등록번호'

    footer_row += 1
    ws[f'A{footer_row}'] = '사업장 소재지'
    ws[f'C{footer_row}'] = '서울특별시 강서구 금낭화로 54-7 (방화동, 동해빌딩 1층)'

    footer_row += 1
    ws[f'A{footer_row}'] = '비고'
    ws[f'C{footer_row}'] = '본 증명서를 위조하거나 변조하는 등 모든 행위에 대한 책임은 판매자에게 있습니다'

    footer_row += 1
    ws[f'A{footer_row}'] = '(주)두라로지스틱스'

    return total_krw


# ── 통화별 수출신고 템플릿 시트 작성 ─────────────────────────────

def write_currency_template_sheet(ws, currency: str,
                                   shopee_data: Optional[dict],
                                   lazada_items: list,
                                   rates: dict,
                                   lazada_write_date: str = ''):
    """
    MYR, PHP, SGD 등 수출신고 프로그램용 시트 작성
    환율: 각 소포수령증 발행일(write_date) 기준
    """
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14

    # 라자다 환율: 큐텐과 동일하게 평균환율 사용
    lazada_rate = rates.get(currency, {}).get('average', 0.0)
    divisor     = RATE_DIVISOR.get(currency, 1)   # VND·JPY → 100, 나머지 → 1

    # ── 쇼피 소계: 각 거래의 발행일 기준 환율 합산 ──
    shopee_fx  = 0.0
    shopee_krw = 0
    if shopee_data:
        for tx in shopee_data.get('transactions', []):
            tx_rate = _get_rate(rates, currency, tx['date'])
            shopee_fx  += tx['amount']
            shopee_krw += round(tx['amount'] * tx_rate / divisor)

    # ── 라자다 소계 ──
    lazada_fx  = sum(it['amount'] for it in lazada_items)
    lazada_krw = round(lazada_fx * lazada_rate / divisor)

    total_krw = shopee_krw + lazada_krw

    # ── 행 1-3 요약 ──
    ws.cell(row=1, column=5, value='쇼피')
    ws.cell(row=1, column=6, value=shopee_fx)
    ws.cell(row=1, column=7, value=shopee_krw)
    ws.cell(row=2, column=5, value='라자다')
    ws.cell(row=2, column=6, value=lazada_fx)
    ws.cell(row=2, column=7, value=lazada_krw)
    ws.cell(row=3, column=7, value=total_krw)

    for row in [1, 2, 3]:
        for col in [5, 6, 7]:
            c = ws.cell(row=row, column=col)
            _style(c, font=FONT_DEFAULT, align=RIGHT)

    # ── 행 4: 헤더 ──
    headers = ['수출신고번호', '기타영세율건수', '선(기)적일자', '통화코드', '환율', '외화금액', '원화금액']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    # ── 행 5+: 쇼피 거래 (각 발행일 기준 환율 개별 적용) ──
    data_row = 5
    transactions = shopee_data.get('transactions', []) if shopee_data else []
    for tx in transactions:
        tx_rate  = _get_rate(rates, currency, tx['date'])
        krw      = round(tx['amount'] * tx_rate / divisor)
        date_int = int(tx['date'].replace('.', '').replace('-', ''))
        row_vals = [None, '1', date_int, currency, tx_rate, tx['amount'], krw]
        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=data_row, column=col, value=v)
            _style(c, font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER)
        data_row += 1

    # ── 라자다 거래 (발행일 기준 환율 적용) ──
    for it in lazada_items:
        krw = round(it['amount'] * lazada_rate / divisor)
        row_vals = [None, '1', None, currency, lazada_rate, it['amount'], krw]
        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=data_row, column=col, value=v)
            _style(c, font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER)
        data_row += 1


# ── 라자다 소포수령증 시트 ───────────────────────────────────────

def write_lazada_receipt_sheet(ws, lazada_data: dict, rates: dict):
    """라자다(소포수령증) 시트"""
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15

    carrier    = lazada_data.get('carrier', '용성종합물류')
    period_end = lazada_data.get('period_end', '')

    # 헤더
    ws['A1'] = (
        f'YONG SUNG LOGISTICS CO., LTD.\n'
        f'ROOM 1215, TOWER A 152, MAGOKSEO-RO, GANGSEO-GU, SEOUL, KOREA\n'
        f'TEL: 82-2-2664-4032  FAX: 82-2-2664-3815\n'
        f'E-mail : admin@yslogic.co.kr    URL : http://www.yslogic.co.kr'
    )
    _style(ws['A1'], font=FONT_DEFAULT, align=LEFT)
    ws.row_dimensions[1].height = 55

    ws['A2'] = '해외화물 소포 수령증'
    _style(ws['A2'], font=FONT_TITLE, align=CENTER)

    ws['A3'] = '1.   제출자 인적 사항'
    _style(ws['A3'], font=FONT_BOLD)

    info_rows = [
        ('사업자등록번호', '529-12-02268', '상호(법인명)', '유엠(UM)'),
        ('성명(대표자)',   '맹진열',       '사업장소재지',
         '서울특별시 광진구 광나루로 556, 1동 2층 2호\n(구의동, 씨엔씨빌딩)'),
        ('거래기간',
         f"{lazada_data.get('period_start','')} – {period_end}",
         '작성일자', lazada_data.get('write_date', '')),
    ]
    for r, (k1, v1, k2, v2) in enumerate(info_rows, 4):
        ws.cell(row=r, column=1, value=k1)
        ws.cell(row=r, column=4, value=v1)
        ws.cell(row=r, column=9, value=k2)
        ws.cell(row=r, column=11, value=v2)

    ws['A7'] = '2.   해외 배송 내역서'
    _style(ws['A7'], font=FONT_BOLD)
    ws['A8'] = '발행사유'
    ws['B8'] = f'{carrier}를 통해 해외로 수출한 내역 증명'

    # 헤더행
    header_row = 9
    for col, h in enumerate(['서비스', '해외배송업체', '출발', '도착', '발송번호', '발송수량', '금액'], 1):
        c = ws.cell(row=header_row, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    # 데이터
    for r, it in enumerate(lazada_data.get('items', []), header_row + 1):
        row_vals = [
            it.get('service', '라자다'),
            it.get('carrier', carrier),
            it.get('origin', 'KR'),
            it.get('destination', ''),
            it.get('tracking_no', ''),
            f"{it.get('qty', '')}건",
            f"{it.get('amount', '')}({it.get('currency', '')})",
        ]
        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            _style(c, font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER)


# ── 큐텐 소포수령증 시트 ────────────────────────────────────────

def write_qoo10_sheet(ws, qoo10_data: Optional[dict], jpy_rate: float):
    """
    큐텐(소포수령증) 시트
    jpy_rate: 거래기간 마지막날 JPY 환율 (100엔 기준)
    """
    ws['A1'] = '해외배송 소포 수령증'
    _style(ws['A1'], font=FONT_TITLE, align=CENTER)

    ws['A3'] = '1.제출자 인적사항'
    _style(ws['A3'], font=FONT_BOLD)

    ws['A5'] = '사업자등록번호'; ws['B5'] = '529-12-02268'
    ws['C5'] = '상호（법인명）'; ws['D5'] = '유엠'
    ws['A6'] = '성명 （대표자）'; ws['B6'] = '맹진열'
    ws['C6'] = '사업장소재지'; ws['D6'] = '서울특별시 광진구 광나루로 556 씨앤씨빌딩 202호 UM'
    ws['A7'] = '거래기간'

    if qoo10_data:
        period = f"{qoo10_data.get('period_start','')} ~ {qoo10_data.get('period_end','')}"
        ws['B7'] = period

    ws['A9'] = '2.해외배송 소포 수령증'
    _style(ws['A9'], font=FONT_BOLD)

    headers = ['판매처', '해외배송업체', '배송국가', '송장번호', '수량', '비고']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=11, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    if qoo10_data:
        ws.cell(row=12, column=1, value='Qoo10')
        ws.cell(row=12, column=2, value='국제로지스틱')
        ws.cell(row=12, column=3, value='일본')
        ws.cell(row=12, column=4, value=qoo10_data.get('tracking_no', ''))
        ws.cell(row=12, column=5, value=f"{qoo10_data.get('qty', '')} 건")

        ws.cell(row=13, column=1, value='당기 해외배송 합계')
        ws.cell(row=13, column=5, value=f"{qoo10_data.get('qty', '')} 건")

        ws['A15'] = '3. 해외배송 내역서'
        _style(ws['A15'], font=FONT_BOLD)
        ws['A17'] = '발행사유'; ws['B17'] = '국제로지스틱을 통해 해외로 수출한 내역 증명'

        detail_headers = ['판매처', '해외배송업체', '출발', '도착', '발송번호', '발송수량', '금액 (JPY)', '원화금액']
        for col, h in enumerate(detail_headers, 1):
            c = ws.cell(row=18, column=col, value=h)
            _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

        jpy_amount = qoo10_data.get('amount', 0)
        krw_amount = round(jpy_amount * jpy_rate / 100)  # JPY는 100엔 기준

        ws.cell(row=18, column=8, value=f'평균환율: {jpy_rate} (100엔)')  # 적용 환율 표시
        ws.cell(row=19, column=1, value='Qoo10')
        ws.cell(row=19, column=2, value='국제로지스틱')
        ws.cell(row=19, column=3, value='KR')
        ws.cell(row=19, column=4, value='JP')
        ws.cell(row=19, column=5, value=qoo10_data.get('tracking_no', ''))
        ws.cell(row=19, column=6, value=f"{qoo10_data.get('qty', '')} 건")
        ws.cell(row=19, column=7, value=jpy_amount)
        ws.cell(row=19, column=8, value=krw_amount)

        ws.cell(row=20, column=1, value='당기 해외배송 합계')
        ws.cell(row=20, column=6, value=f"{qoo10_data.get('qty', '')} 건")
        ws.cell(row=20, column=7, value=jpy_amount)
        ws.cell(row=20, column=8, value=krw_amount)
    else:
        ws['A12'] = '⚠️ 큐텐 데이터 없음 — STEP 2에서 수동 입력하세요'
        _style(ws['A12'], font=Font(name='맑은 고딕', size=9, color='FF0000'))


# ── 총집계 시트 ─────────────────────────────────────────────────

def write_summary_sheet(ws, shopee_totals: dict, lazada_totals: dict,
                         qoo10_data: Optional[dict], jpy_rate: float,
                         year_month: str):
    """총집계 시트 작성"""
    ws['A1'] = '유엠(UM)(529-12-02268)'
    _style(ws['A1'], font=FONT_TITLE)

    ws['B2'] = year_month  # 예: '2025년 12월'
    _style(ws['B2'], font=FONT_BOLD)

    # 쇼피 소계
    ws['G4'] = '쇼피'
    _style(ws['G4'], font=FONT_BOLD, fill=SUBHEAD_FILL, align=CENTER)
    ws['G13'] = '국가'
    ws['H13'] = '외화'
    ws['I13'] = '원화'
    _style(ws['G13'], font=FONT_BOLD, align=CENTER, border=THIN_BORDER)
    _style(ws['H13'], font=FONT_BOLD, align=CENTER, border=THIN_BORDER)
    _style(ws['I13'], font=FONT_BOLD, align=CENTER, border=THIN_BORDER)

    COUNTRY_NAMES = {
        'MYR': '말레이시아(MYR)', 'PHP': '필리핀(PHP)',
        'SGD': '싱가폴(SGD)', 'THB': '태국(THB)',
        'TWD': '대만(TWD)', 'VND': '베트남(VND)',
    }

    shopee_total_krw = 0
    for r, (cur, name) in enumerate(COUNTRY_NAMES.items(), 14):
        data = shopee_totals.get(cur, {})
        fx  = data.get('fx', 0.0)
        krw = data.get('krw', 0)
        shopee_total_krw += krw
        ws.cell(row=r, column=7, value=name)
        ws.cell(row=r, column=8, value=fx)
        ws.cell(row=r, column=9, value=krw)

    ws.cell(row=20, column=7, value='총합')
    ws.cell(row=20, column=9, value=shopee_total_krw)
    _style(ws.cell(row=20, column=7), font=FONT_BOLD)

    # 라자다 소계
    ws['G22'] = '라자다'
    _style(ws['G22'], font=FONT_BOLD, fill=SUBHEAD_FILL, align=CENTER)
    ws['G23'] = '국가'; ws['H23'] = '외화'; ws['I23'] = '원화'

    lazada_total_krw = 0
    LAZADA_COUNTRIES = ['MYR', 'PHP', 'SGD', 'VND']
    for r, cur in enumerate(LAZADA_COUNTRIES, 24):
        data = lazada_totals.get(cur, {})
        fx  = data.get('fx', 0.0)
        krw = data.get('krw', 0)
        lazada_total_krw += krw
        ws.cell(row=r, column=7, value=COUNTRY_NAMES.get(cur, cur))
        ws.cell(row=r, column=8, value=fx)
        ws.cell(row=r, column=9, value=krw)

    ws.cell(row=28, column=7, value='총합')
    ws.cell(row=28, column=9, value=lazada_total_krw)
    _style(ws.cell(row=28, column=7), font=FONT_BOLD)

    # 큐텐
    ws['G30'] = '큐텐'
    _style(ws['G30'], font=FONT_BOLD, fill=SUBHEAD_FILL, align=CENTER)
    ws['G31'] = '외화'; ws['H31'] = '평균환율'; ws['I31'] = '원화'

    if qoo10_data:
        jpy_amount = qoo10_data.get('amount', 0)
        krw = round(jpy_amount * jpy_rate / 100)
        ws.cell(row=32, column=7, value=jpy_amount)
        ws.cell(row=32, column=8, value=jpy_rate)
        ws.cell(row=32, column=9, value=krw)


# ── 전체 엑셀 생성 ───────────────────────────────────────────────

def generate_excel(
    shopee_results: list,      # [parse_shopee_pdf() 결과, ...]
    lazada_result:  Optional[dict],   # parse_lazada_pdf() 결과
    qoo10_result:   Optional[dict],   # parse_qoo10_pdf() 결과
    rates:          dict,      # fetch_all_currencies() 결과
    output_path:    str,
    year:           int,
    month:          int,
):
    """전체 엑셀 파일 생성"""
    wb = Workbook()
    wb.remove(wb.active)

    currency_list = ['MYR', 'PHP', 'SGD', 'THB', 'TWD', 'VND']

    # ── 라자다 발행일 추출 (write_date → period_end fallback) ──
    if lazada_result:
        lazada_write_date = (lazada_result.get('write_date', '')
                             or lazada_result.get('period_end', ''))
    else:
        lazada_write_date = ''

    # ── 큐텐 JPY 환율: 거래 전체가 한 달에 걸쳐 있으므로 평균환율 사용 ──
    # (개별 거래 날짜가 없어 발행일/기간말 기준이 아닌 월 평균이 정확함)
    jpy_rate_data = rates.get('JPY')
    if jpy_rate_data:
        jpy_rate = jpy_rate_data.get('average', 0.0)
        # average가 0이면 일별 평균 직접 계산
        if jpy_rate == 0.0:
            daily = jpy_rate_data.get('daily', [])
            if daily:
                jpy_rate = round(sum(d['rate'] for d in daily) / len(daily), 2)
    else:
        jpy_rate = 0.0
    # write_date 보존 (선적일자 기재용)
    qoo10_write_date = ''
    if qoo10_result:
        qoo10_write_date = (qoo10_result.get('write_date', '')
                            or qoo10_result.get('period_end', ''))

    # ── 총집계 ──────────────────────────────────────────────
    ws_summary = wb.create_sheet('총집계')
    shopee_totals = {}
    lazada_totals = {}

    for sd in shopee_results:
        cur = sd.get('currency', '')
        if not cur:
            continue
        # 각 거래의 발행일(tx['date']) 기준 환율로 개별 계산 후 합산
        div = RATE_DIVISOR.get(cur, 1)
        total_fx  = 0.0
        total_krw = 0
        for tx in sd.get('transactions', []):
            tx_rate    = _get_rate(rates, cur, tx['date'])
            total_fx  += tx['amount']
            total_krw += round(tx['amount'] * tx_rate / div)
        # 거래 내역 없으면 total_amount 사용 (fallback)
        if not sd.get('transactions'):
            rate_date = sd.get('write_date', '') or sd.get('period_end', '')
            rate = _get_rate(rates, cur, rate_date)
            total_fx  = sd.get('total_amount', 0.0)
            total_krw = round(total_fx * rate / div)
        shopee_totals[cur] = {'fx': total_fx, 'krw': total_krw}

    if lazada_result:
        laz_rate_by_cur = {}
        for it in lazada_result.get('items', []):
            cur = it.get('currency', '')
            if cur not in laz_rate_by_cur:
                laz_rate_by_cur[cur] = rates.get(cur, {}).get('average', 0.0)
            rate = laz_rate_by_cur[cur]
            div  = RATE_DIVISOR.get(cur, 1)
            krw  = round(it.get('amount', 0.0) * rate / div)
            if cur not in lazada_totals:
                lazada_totals[cur] = {'fx': 0.0, 'krw': 0}
            lazada_totals[cur]['fx']  += it.get('amount', 0.0)
            lazada_totals[cur]['krw'] += krw

    write_summary_sheet(ws_summary, shopee_totals, lazada_totals,
                        qoo10_result, jpy_rate,
                        f'{year}년 {month:02d}월')

    # ── 통화별 수출신고 템플릿 (MYR, PHP, SGD, THB, TWD, VND) ──
    for cur in currency_list:
        ws = wb.create_sheet(cur)
        sd = next((s for s in shopee_results if s.get('currency') == cur), None)
        lazada_items = []
        if lazada_result:
            lazada_items = [it for it in lazada_result.get('items', [])
                            if it.get('currency') == cur]
        write_currency_template_sheet(ws, cur, sd, lazada_items, rates,
                                      lazada_write_date=lazada_write_date)

    # ── JPY 수출신고 시트 (큐텐만) ──
    ws_jpy = wb.create_sheet('JPY')
    headers = ['수출신고번호', '기타영세율건수', '선(기)적일자', '통화코드', '환율', '외화금액', '원화금액']
    for col, h in enumerate(headers, 1):
        c = ws_jpy.cell(row=4, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)
    if qoo10_result:
        jpy_amount = qoo10_result.get('amount', 0)
        krw = round(jpy_amount * jpy_rate / 100)
        ws_jpy.cell(row=1, column=5, value='큐텐')
        ws_jpy.cell(row=1, column=6, value=jpy_amount)
        ws_jpy.cell(row=1, column=7, value=krw)
        # 데이터행 — 선적일자에 발행일 사용
        date_str = ''
        if qoo10_write_date:
            date_str = int(qoo10_write_date.replace('-', ''))
        ws_jpy.cell(row=5, column=3, value=date_str or None)
        ws_jpy.cell(row=5, column=4, value='JPY')
        ws_jpy.cell(row=5, column=5, value=jpy_rate)
        ws_jpy.cell(row=5, column=6, value=jpy_amount)
        ws_jpy.cell(row=5, column=7, value=krw)

    # ── 큐텐(소포수령증) ──
    ws_q10 = wb.create_sheet('큐텐(소포수령증)')
    write_qoo10_sheet(ws_q10, qoo10_result, jpy_rate)

    # ── 쇼피 국가별 시트 ──
    shopee_sheet_names = {
        'MYR': '쇼피(MYR)', 'PHP': '쇼피(PHP)', 'SGD': '쇼피(SGD)',
        'THB': '쇼피(THB)', 'TWD': '쇼피(TWD)', 'VND': '쇼피(VND)',
    }
    for cur, sheet_name in shopee_sheet_names.items():
        ws = wb.create_sheet(sheet_name)
        sd = next((s for s in shopee_results if s.get('currency') == cur), None)
        if sd:
            write_shopee_sheet(ws, sd, rates)
        else:
            ws['A1'] = f'{sheet_name} 데이터 없음'

    # ── 라자다(소포수령증) + 라자다(국가별) ──
    ws_laz = wb.create_sheet('라자다(소포수령증)')
    if lazada_result:
        write_lazada_receipt_sheet(ws_laz, lazada_result, rates)
    else:
        ws_laz['A1'] = '라자다 데이터 없음'

    for cur in ['MYR', 'PHP', 'SGD', 'VND']:
        ws = wb.create_sheet(f'라자다({cur})')
        if lazada_result:
            items = [it for it in lazada_result.get('items', []) if it.get('currency') == cur]
            if items:
                headers = ['No', 'OBD DT', 'HBL No', 'MBL No', 'POL', 'POD', 'PKG', 'PKG Unit', 'G.WT', 'C.WT']
                for col, h in enumerate(headers, 1):
                    c = ws.cell(row=2, column=col, value=h)
                    _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    # ── 환율 시트 ──
    all_currencies = ['JPY', 'TWD', 'THB', 'SGD', 'MYR', 'PHP', 'VND', 'BRL']
    for cur in all_currencies:
        ws = wb.create_sheet(f'환율({cur})')
        write_exchange_rate_sheet(ws, rates.get(cur))

    wb.save(output_path)
    print(f'  ✅ 엑셀 저장 완료: {output_path}')
    return output_path
